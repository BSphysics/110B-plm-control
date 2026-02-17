
import pandas as pd
import numpy as np
from tkinter import Tk, filedialog
from generatePhaseTilt import generate_phase_tilt
from HGMode import HG_mode
from ampModPhase import amp_mod_phase
import os
import time
from numba import njit
from datetime import datetime
from pypylon import pylon
from moveSliderToNotAttenuator import move_slider_to_not_attenuator
from polMeasure import pol_measure
from polAnalyse49Beams import pol_analyse_49_beams
from simpleBeamMaker import simple_beam_maker

@njit
def add_beams(acc, A_amp, A_phase, B_amp, B_phase, global_amp):
    rows, cols = A_amp.shape
    for i in range(rows):
        for j in range(cols):
            a = A_amp[i, j] * np.cos(A_phase[i, j]) + 1j * A_amp[i, j] * np.sin(A_phase[i, j])
            b = B_amp[i, j] * np.cos(B_phase[i, j]) + 1j * B_amp[i, j] * np.sin(B_phase[i, j])
            acc[i, j] += global_amp * (a + b)

def beam_A_phase_optimiser_49_beams(self, plm, camera):
    plm.pause_ui()
# Hide the root tkinter window
    root = Tk()
    root.withdraw()

    # Set a default path (relative or absolute)
    default_path = os.path.join('D:\PLM\plm python control\wrappers', "multibeam parameters")

    # Open file dialog with default path
    xlsx_file_path = filedialog.askopenfilename(
        title="Select the spreadsheet",
        filetypes=[("Excel files", "*.xlsx *.xls")],
        initialdir=default_path
    )

    if not xlsx_file_path:
        print("No file selected.")
        return None  # Or raise an exception / return default values

    df = pd.read_excel(xlsx_file_path) 
    beamParameters = df.iloc[:, 3].values #.dropna().tolist() 

    beamParameterBlocks = []
    for i in range(0, len(beamParameters), 11):  
        chunk = beamParameters[i:i+10]
        if len(chunk) == 10 and not np.any(pd.isna(chunk)):
            beamParameterBlocks.append(chunk)

    beamParameterBlocks = np.array(beamParameterBlocks)

    cols, rows = 1358 , 800

    N = 1358
    M = 800
    numHolograms = 24
    
    combinedComplexs = []

    beamA_HG_phase, beamA_HG_amplitude = HG_mode(cols, rows, self.user_values[4], self.user_values[5], self.user_values[8], self.user_values[9], self.user_values[12], self.user_values[13])
    beamB_HG_phase, beamB_HG_amplitude = HG_mode(cols, rows, self.user_values[6], self.user_values[7], self.user_values[10], self.user_values[11], self.user_values[14], self.user_values[15])
    if self.clear_beam_A_correction_flag == True:
        self.beam_A_correction_data = np.zeros_like(beamA_phase_tilt) # clear the correction

    if self.clear_beam_B_correction_flag == True:
        self.beam_B_correction_data = np.zeros_like(beamB_phase_tilt)

    finalCombined = np.zeros((rows, cols), dtype=np.complex128)
    for i, chunk in enumerate(beamParameterBlocks):
        print('Loading parameters for Beam: ' + str(int(i)))

        beamA_phase_tilt = generate_phase_tilt(rows, cols, chunk[0], chunk[1], self.button_states[0], self.button_states[1]) 
        beamB_phase_tilt = generate_phase_tilt(rows, cols, chunk[2], chunk[3], self.button_states[2], self.button_states[3])

        relative_amplitudes = np.array([chunk[7],chunk[8]])

        max_val = np.max(relative_amplitudes)
        if max_val > 0:
            relative_amplitudes = relative_amplitudes / max_val

        beamA_amplitude = beamA_HG_amplitude * relative_amplitudes[0]
        beamB_amplitude = beamB_HG_amplitude * relative_amplitudes[1]
        global_amplitude = chunk[9]
        if abs(global_amplitude) < 1e-6:
            continue
        
        relative_phase_A = (chunk[4]/100)*2*np.pi  
        relative_phase_B = (chunk[5]/100)*2*np.pi
        #print('Relative Phase Beam A = ' + str(np.round(relative_phase_A,2)))

        global_phase = (chunk[6]/100)*2*np.pi

        beamA_phase = beamA_phase_tilt - self.beam_A_correction_data + beamA_HG_phase + relative_phase_A + global_phase
        beamB_phase = beamB_phase_tilt - self.beam_B_correction_data + beamB_HG_phase + relative_phase_B + global_phase

        add_beams(finalCombined, beamA_amplitude, beamA_phase, beamB_amplitude, beamB_phase, global_amplitude)
        finalCombined[np.abs(finalCombined) < 1e-6] = 0

    amplitude_modulated_combined_phase = amp_mod_phase(finalCombined) 
    plm_phase_map = (amplitude_modulated_combined_phase + np.pi) / (2*np.pi)
    plm_frame = np.broadcast_to(plm_phase_map[:, :, None], (M, N, numHolograms)).astype(np.float32)
    plm_frame = np.transpose(plm_frame, (1, 0, 2)).copy(order='F')
                                
    plm.bitpack_and_insert_gpu(plm_frame, 1)
    plm.resume_ui()
    plm.set_frame(1)
    time.sleep(0.2)
    plm.play()
    plm.play()
    move_slider_to_not_attenuator(self)

    global_amplitudes = np.array([self.user_values[16] , self.user_values[17]])
    global_amplitudes = global_amplitudes/ np.max(global_amplitudes)

    file_path = pol_measure(self.camera, global_amplitudes, '_Rotating linear polariser_' , 400)
    print(file_path)
    mean_intensity_ellipticity = pol_analyse_49_beams(manual_path=file_path)

    ##########################################################################

    
    beamAPhaseShifts=[]
    mean_intensity_ellipticities = []
    for beamAPhaseShift in range(-15, 16, 2):   # START → STOP in steps of STEP
        beamAPhaseShifts.append(beamAPhaseShift)
        print('Globally shifting phase of Beam A by ' + str(beamAPhaseShift) + '%')
        plm.pause_ui()

        finalCombined = np.zeros((rows, cols), dtype=np.complex128)
        for i, chunk in enumerate(beamParameterBlocks):
            print('Loading parameters for Beam: ' + str(int(i)))

            beamA_phase_tilt = generate_phase_tilt(rows, cols, chunk[0], chunk[1], self.button_states[0], self.button_states[1]) 
            beamB_phase_tilt = generate_phase_tilt(rows, cols, chunk[2], chunk[3], self.button_states[2], self.button_states[3])

            relative_amplitudes = np.array([chunk[7],chunk[8]])

            max_val = np.max(relative_amplitudes)
            if max_val > 0:
                relative_amplitudes = relative_amplitudes / max_val

            beamA_amplitude = beamA_HG_amplitude * relative_amplitudes[0]
            beamB_amplitude = beamB_HG_amplitude * relative_amplitudes[1]
            global_amplitude = chunk[9]
            if abs(global_amplitude) < 1e-6:
                continue
        
            relative_phase_A = ((chunk[4] + beamAPhaseShift)/100)*2*np.pi  
            relative_phase_B = (chunk[5]/100)*2*np.pi
            #print('Relative Phase Beam A = ' + str(np.round(relative_phase_A,2)))

            global_phase = (chunk[6]/100)*2*np.pi

            beamA_phase = beamA_phase_tilt - self.beam_A_correction_data + beamA_HG_phase + relative_phase_A + global_phase
            beamB_phase = beamB_phase_tilt - self.beam_B_correction_data + beamB_HG_phase + relative_phase_B + global_phase

            add_beams(finalCombined, beamA_amplitude, beamA_phase, beamB_amplitude, beamB_phase, global_amplitude)
            finalCombined[np.abs(finalCombined) < 1e-6] = 0


        amplitude_modulated_combined_phase = amp_mod_phase(finalCombined) 
        plm_phase_map = (amplitude_modulated_combined_phase + np.pi) / (2*np.pi)
        self.multibeam_flag = False
        plm_frame = np.broadcast_to(plm_phase_map[:, :, None], (M, N, numHolograms)).astype(np.float32)
        plm_frame = np.transpose(plm_frame, (1, 0, 2)).copy(order='F')
                                    
        plm.bitpack_and_insert_gpu(plm_frame, 1)
        plm.resume_ui()
        plm.set_frame(1)
        time.sleep(0.2)
        plm.play()
        plm.play()

        file_path = pol_measure(self.camera, global_amplitudes, '_Rotating linear polariser_' , 400)
        mean_intensity_ellipticity = pol_analyse_49_beams(manual_path=file_path)[0]
        mean_intensity_ellipticities.append(mean_intensity_ellipticity)


    print("\nPhase Shift vs Mean Intensity Ellipticity")
    print("-" * 50)
    print(f"{'Phase Shift (%)':>18} | {'Mean Ellipticity':>20}")
    print("-" * 50)

    for shift, ellip in zip(beamAPhaseShifts, mean_intensity_ellipticities):
        print(f"{shift:>18} | {ellip:>20.6f}")

    print("-" * 50)

    # ---- Find optimal shift ----
    min_index = np.argmin(mean_intensity_ellipticities)
    optimal_shift = beamAPhaseShifts[min_index]

    from openpyxl import load_workbook
    wb = load_workbook(xlsx_file_path)
    ws = wb.active  # This targets the currently active sheet
    current_excel_row = 2

    for idx in range(49):
        beamParameterBlocks[idx][4] = beamParameterBlocks[idx][4] + optimal_shift  

    for chunk in beamParameterBlocks:
        # chunk is your 10 values for one beam
        for value in chunk:
            # column=4 is Column D
            ws.cell(row=current_excel_row, column=4).value = value
            current_excel_row += 1
        current_excel_row += 1

    folderName = os.path.dirname(xlsx_file_path)
    fileName = os.path.basename(xlsx_file_path)
    name, ext = os.path.splitext(fileName)  # split name and .xlsx

    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")

    xlsx_save_path = os.path.join(
        folderName,
        f"{name}_BeamAPhaseShift_{optimal_shift}_{timestamp}{ext}"
    )

    # Save workbook
    wb.save(xlsx_save_path)
    print(f"Saved optimised file to: {xlsx_save_path}")

