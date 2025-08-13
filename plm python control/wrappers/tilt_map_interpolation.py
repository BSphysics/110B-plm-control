# -*- coding: utf-8 -*-
"""
Created on Mon Jun 16 13:24:05 2025

@author: bs426
"""


import numpy as np
from matplotlib import pyplot as plt

from scipy.interpolate import LinearNDInterpolator
import numpy as np

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

import os
from datetime import datetime

folder_name = r'D:\PLM\plm python control\wrappers\Data\2025_06_18\15_21_07 tilt_mapping Beam B'


# Load your data
centroid_data = np.load(os.path.join(folder_name , 'tilt_map_data.npy'))
flat_data = centroid_data.reshape(-1, 4)
x_input, y_input, centroid_x, centroid_y = flat_data.T

print('\nSmallest accessible x pixel coord on camera = ' + str(np.min(centroid_x)))
print('\nSmallest accessible y pixel coord on camera = ' + str(np.min(centroid_y)))

print('\nLargest accessible x pixel coord on camera = ' + str(np.max(centroid_x)))
print('\nLargest accessible y pixel coord on camera = ' + str(np.max(centroid_y)))
# Build interpolators
x_interp = LinearNDInterpolator(list(zip(centroid_x, centroid_y)), x_input)
y_interp = LinearNDInterpolator(list(zip(centroid_x, centroid_y)), y_input)

# Function to get input (x, y) from desired camera position
def get_xy_from_camera_target(centroid_x_target, centroid_y_target):
    x_cmd = x_interp(centroid_x_target, centroid_y_target)
    y_cmd = y_interp(centroid_x_target, centroid_y_target)
    return x_cmd, y_cmd

# Example usage
target_camera_pos = (250, 250)
x_cmd, y_cmd = get_xy_from_camera_target(*target_camera_pos)



grid_x = np.linspace(100, 240, 10)
grid_y = np.linspace(160, 300, 10)

# Create 2D grid of 10 by 10 camera target positions
camera_target_grid = np.array([[cx, cy] for cx in grid_x for cy in grid_y])

# Interpolate to get x/y control inputs
x_cmds = x_interp(camera_target_grid[:, 0], camera_target_grid[:, 1])
y_cmds = y_interp(camera_target_grid[:, 0], camera_target_grid[:, 1])

# Combine into final array
camera_targets_with_commands = np.column_stack((camera_target_grid, x_cmds, y_cmds))

commands = camera_targets_with_commands[:, 2:]  # shape (25, 2)

# Try to load existing workbook, or create a new one
now = datetime.now()
filename = now.strftime("%H_%M_%S") + ' multiBeamData.xlsx'

full_filename = os.path.join(folder_name , filename)

try:
    wb = load_workbook(full_filename)
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active

# Write commands into column D with 11-row spacing
for i, (x_cmd, y_cmd) in enumerate(commands):
    if x_cmd is None or not np.isfinite(x_cmd):
        x_cmd = 500

    if y_cmd is None or not np.isfinite(y_cmd):
        y_cmd = 500

    base_row = 11 * i + 2  # Start of each block
    for j in range(10):  # 10 rows per block
        row = base_row + j
        ws[f"B{row}"] = (i+1)
    
    BeamA_x_cell = f"D{11 * i + 2}"  # Row for x
    BeamA_y_cell = f"D{11 * i + 3}"  # Row for y
    ws[BeamA_x_cell] = x_cmd
    ws[BeamA_y_cell] = y_cmd
    
    BeamB_x_cell = f"D{11 * i + 5}"  # Row for x
    BeamB_y_cell = f"D{11 * i + 4}"  # Row for y
    ws[BeamB_x_cell] = 30
    ws[BeamB_y_cell] = 12
    
    BeamA_rel_phase = f"D{11 * i + 6}"  
    BeamB_rel_phase = f"D{11 * i + 7}"  
    ws[BeamA_rel_phase] = 60
    ws[BeamB_rel_phase] = 0
    
    global_phase = f"D{11 * i + 8}"  
    ws[global_phase] = 0
    
    BeamA_rel_amplitude = f"D{11 * i + 9}"  
    BeamB_rel_amplitude = f"D{11 * i + 10}"  
    ws[BeamA_rel_amplitude] = 1
    ws[BeamB_rel_amplitude] = 0
    
    global_amplitude = f"D{11 * i + 11}"  
    ws[global_amplitude] = 1
    
    BeamA_label = f"A{11 * i + 2}"
    ws[BeamA_label] = 'Beam A'
    BeamA_label = f"A{11 * i + 3}"
    ws[BeamA_label] = 'Beam A'
    
    tiltX_label = f"C{11 * i + 2}"
    ws[tiltX_label] = 'Grating period X (plm pix)'
    tiltY_label = f"C{11 * i + 3}"
    ws[tiltY_label] = 'Grating period Y (plm pix)'
    
    BeamB_label = f"A{11 * i + 4}"
    ws[BeamB_label] = 'Beam B'
    BeamB_label = f"A{11 * i + 5}"
    ws[BeamB_label] = 'Beam B'
    
    tiltX_label = f"C{11 * i + 4}"
    ws[tiltX_label] = 'Grating period X (plm pix)'
    tiltY_label = f"C{11 * i + 5}"
    ws[tiltY_label] = 'Grating period Y (plm pix)'
    
    BeamA_label = f"A{11 * i + 6}"
    ws[BeamA_label] = 'Beam A'
    phase_label = f"C{11 * i + 6}"
    ws[phase_label] = 'Relative phase (%)'
    
    BeamB_label = f"A{11 * i + 7}"
    ws[BeamB_label] = 'Beam B'
    phase_label = f"C{11 * i + 7}"
    ws[phase_label] = 'Relative phase (%)'
    
    Beam_label = f"A{11 * i + 8}"
    ws[Beam_label] = 'Beam'
    phase_label = f"C{11 * i + 8}"
    ws[phase_label] = 'Global phase (%)'
    
    BeamA_label = f"A{11 * i + 9}"
    ws[BeamA_label] = 'Beam A'
    amp_label = f"C{11 * i + 9}"
    ws[amp_label] = 'Relative amplitude'
    
    BeamB_label = f"A{11 * i + 10}"
    ws[BeamB_label] = 'Beam B'
    amp_label = f"C{11 * i + 10}"
    ws[amp_label] = 'Relative amplitude'
    
    Beam_label = f"A{11 * i + 11}"
    ws[Beam_label] = 'Beam'
    amp_label = f"C{11 * i + 11}"
    ws[amp_label] = 'Global amplitude'
    
   

# Save the workbook
wb.save(full_filename)
print(f"\n Wrote {len(commands)} x-y command pairs to '{full_filename}'")