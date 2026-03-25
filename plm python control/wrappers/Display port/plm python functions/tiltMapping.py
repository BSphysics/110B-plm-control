import pandas as pd
import numpy as np
from generatePhaseTilt import generate_phase_tilt
from HGMode import HG_mode
from ampModPhase import amp_mod_phase
from pypylon import pylon
import os
from datetime import datetime
import time
from basler_centroid import baslerCentroid
import cv2
from scipy.interpolate import LinearNDInterpolator
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import sys
from PyQt5.QtWidgets import QApplication, QInputDialog
import matplotlib.pyplot as plt

numHolograms = 24

def tilt_mapping(self, plm, camera):
    cols, rows = 1358 , 800

    beamName, ok = QInputDialog.getText(None, "Select Beam", 'Which beam are you tilt mapping?\nType "a" for Beam A, or "b" for Beam B')

    # If user presses OK and enters a valid input
    if ok:
        beamName = beamName.strip().lower()

        if beamName == 'a':
            tiltx = np.linspace(10, 14, 10)
            tilty = np.linspace(10, 15, 10)
            beam_name = 'Beam A'

        elif beamName == 'b':
            #tiltx = np.linspace(24, 90, 10)
            s = np.linspace(0, 1, 10)  # 10 samples
            s_skewed = s**2  # or try **3 for stronger skew

            # Map back to the tilt range [24, 90]
            tiltx = 24 + s_skewed * (90 - 24)
            tilty = np.linspace(10, 15, 10)
            beam_name = 'Beam B'

        else:
            print("Invalid beam selection. Please enter 'a' or 'b'.")
            sys.exit(1)
    else:
        print("User cancelled input.")
        sys.exit(1)

    total_patterns = len(tiltx) * len(tilty)
    pattern_counter = 0

    # Create folder with today's date
    now = datetime.now()
    date_str = now.strftime("%Y_%m_%d")
    date_folder = os.path.join(os.getcwd(), 'Data', date_str)
    os.makedirs(date_folder, exist_ok=True)

    # Create subfolder with current time inside the date folder
    time_str = now.strftime("%H_%M_%S") + ' tilt_mapping ' + beam_name
    time_folder = os.path.join(date_folder, time_str)
    os.makedirs(time_folder, exist_ok=True)
    idx=0

    mapping_data = np.zeros((tiltx.size, tilty.size, 4))

    for i, x in enumerate(tiltx):
        for j, y in enumerate(tilty):

            plm.pause_ui()
            pattern_counter += 1
            print(f"plm pattern {pattern_counter}/{total_patterns}")
            idx+=1

            if beamName == 'a':
                beamA_phase_tilt = generate_phase_tilt(rows, cols, x, y, self.button_states[0], self.button_states[1]) 
                beamA_HG_phase, beamA_HG_amplitude = HG_mode(cols, rows, self.user_values[4], self.user_values[5], self.user_values[8], self.user_values[9], self.user_values[12], self.user_values[13])
                beamA_phase = beamA_phase_tilt - self.beam_A_correction_data + beamA_HG_phase
                beamA_amplitude = 1
                beamAcomplex = beamA_amplitude * np.exp(1j * beamA_phase) 
                amplitude_modulated_combined_phase = amp_mod_phase(beamAcomplex)

            if beamName == 'b':
                beamB_phase_tilt = generate_phase_tilt(rows, cols, x, y, self.button_states[2], self.button_states[3]) 
                beamB_HG_phase, beamB_HG_amplitude = HG_mode(cols, rows, self.user_values[6], self.user_values[7], self.user_values[10], self.user_values[11], self.user_values[14], self.user_values[15])
                beamB_phase = beamB_phase_tilt - self.beam_B_correction_data + beamB_HG_phase
                beamB_amplitude = 1
                beamBcomplex = beamB_amplitude * np.exp(1j * beamB_phase) 
                amplitude_modulated_combined_phase = amp_mod_phase(beamBcomplex)

            plm_phase_map = (amplitude_modulated_combined_phase + np.pi) / (2*np.pi)

            plm_frame = np.broadcast_to(plm_phase_map[:, :, None], (rows, cols, numHolograms)).astype(np.float32)
            plm_frame = np.transpose(plm_frame, (1, 0, 2)).copy(order='F')
                                        
            plm.bitpack_and_insert_gpu(plm_frame, 1)
            plm.resume_ui()

            plm.set_frame(1)
            
            time.sleep(0.2)
            plm.play()
            plm.play()

            if camera.IsGrabbing():
                camera.StopGrabbing()

            images_per_batch = 10

            # Set to software trigger mode
            camera.TriggerSelector.SetValue('FrameStart')
            camera.TriggerMode.SetValue('On')
            camera.TriggerSource.SetValue('Software')
            camera.ExposureTimeAbs.SetValue(200)

            # Start grabbing manually
            camera.StartGrabbingMax(images_per_batch)

            # Get image dimensions
            camera_width = camera.Width.Value
            camera_height = camera.Height.Value

            # Pre-allocate array: shape [images_per_batch, height, width]
            all_images = np.zeros((images_per_batch, camera_height, camera_width), dtype=np.uint8)

            grab_result = None

            for image_index in range(images_per_batch):
                # Issue software trigger
                camera.TriggerSoftware.Execute()

                # Wait for image
                grab_result = camera.RetrieveResult(5000, pylon.TimeoutHandling_ThrowException)

                if grab_result.GrabSucceeded():
                    img = grab_result.Array
                    all_images[image_index, :, :] = img
                else:
                    print("Failed to grab image", image_index)

                grab_result.Release()


            self.centroid_x, self.centroid_y = baslerCentroid(np.mean(all_images, axis=0), 3, 5)
            #print('XY centroid = ' + str(self.centroid_x) + ' ' + str(self.centroid_y) + ' pix')
            mapping_data[i, j] = [x, y, self.centroid_x, self.centroid_y]
    
            color = (0, 0, 255)  # BGR format for red (OpenCV)
            radius = 10  # You can adjust the size of the dot if needed
            thickness = 2  # Fill the circle
            img_bgr = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
            img_with_dot = cv2.circle(img_bgr.copy(), (self.centroid_x, self.centroid_y), radius, color, thickness)
            filename = os.path.join(time_folder, f"image_{idx:03d}_dot.png") 
            cv2.imwrite(filename, img_with_dot) 


            

    np.save(os.path.join(time_folder, beam_name + ' tilt_map_data.npy'), mapping_data)
    #centroid_data = np.load('tilt_map_data.npy')
    flat_data = mapping_data.reshape(-1, 4)
    x_input, y_input, centroid_x, centroid_y = flat_data.T

    print('\nSmallest accessible x pixel coord on camera = ' + str(np.min(centroid_x)))
    print('\nSmallest accessible y pixel coord on camera = ' + str(np.min(centroid_y)))

    print('\nLargest accessible x pixel coord on camera = ' + str(np.max(centroid_x)))
    print('\nLargest accessible y pixel coord on camera = ' + str(np.max(centroid_y)))


    # Build interpolators
    x_interp = LinearNDInterpolator(list(zip(centroid_x, centroid_y)), x_input)
    y_interp = LinearNDInterpolator(list(zip(centroid_x, centroid_y)), y_input)

    # Function to get input (x, y) from desired camera position
    def get_xy_from_camera_target(centroid_x_target, centroid_y_target):
        x_cmd = x_interp(centroid_x_target, centroid_y_target)
        y_cmd = y_interp(centroid_x_target, centroid_y_target)
        return x_cmd, y_cmd

    grid_x = np.linspace(180, 260, 7)
    grid_y = np.linspace(210, 290, 7)

    # Create 2D grid of 10 by 10 camera target positions
    camera_target_grid = np.array([[cx, cy] for cx in grid_x for cy in grid_y])

    # Interpolate to get x/y control inputs
    x_cmds = x_interp(camera_target_grid[:, 0], camera_target_grid[:, 1])
    y_cmds = y_interp(camera_target_grid[:, 0], camera_target_grid[:, 1])

    # Combine into final array
    camera_targets_with_commands = np.column_stack((camera_target_grid, x_cmds, y_cmds))

    commands = camera_targets_with_commands[:, 2:]  # shape (25, 2)

    # Try to load existing workbook, or create a new one
    now = datetime.now()
    filename = now.strftime("%H_%M_%S__") + beam_name + ' multiBeamData.xlsx'
    full_filename = os.path.join(time_folder , filename)
    try:
        wb = load_workbook(full_filename)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active

    # Write commands into column D with 11-row spacing
    for i, (x_cmd, y_cmd) in enumerate(commands):
        if x_cmd is None or not np.isfinite(x_cmd):
            x_cmd = 500

        if y_cmd is None or not np.isfinite(y_cmd):
            y_cmd = 500

        base_row = 11 * i + 2  # Start of each block
        for j in range(10):  # 10 rows per block
            row = base_row + j
            ws[f"B{row}"] = (i+1)
        
        if beamName == 'a': 
            BeamA_x_cell = f"D{11 * i + 2}"  # Row for x
            BeamA_y_cell = f"D{11 * i + 3}"  # Row for y
            ws[BeamA_x_cell] = x_cmd
            ws[BeamA_y_cell] = y_cmd
            BeamB_x_cell = f"D{11 * i + 4}"  # Row for x
            BeamB_y_cell = f"D{11 * i + 5}"  # Row for y
            ws[BeamB_x_cell] = 30
            ws[BeamB_y_cell] = 12
        
        elif beamName == 'b': 
            BeamA_x_cell = f"D{11 * i + 2}"  # Row for x
            BeamA_y_cell = f"D{11 * i + 3}"  # Row for y
            ws[BeamA_x_cell] = 10
            ws[BeamA_y_cell] = 12
            BeamB_x_cell = f"D{11 * i + 4}"  # Row for x
            BeamB_y_cell = f"D{11 * i + 5}"  # Row for y
            ws[BeamB_x_cell] = x_cmd
            ws[BeamB_y_cell] = y_cmd

        BeamA_rel_phase = f"D{11 * i + 6}"  
        BeamB_rel_phase = f"D{11 * i + 7}"  
        ws[BeamA_rel_phase] = 60
        ws[BeamB_rel_phase] = 0
        
        global_phase = f"D{11 * i + 8}"  
        ws[global_phase] = 0
        
        BeamA_rel_amplitude = f"D{11 * i + 9}"  
        BeamB_rel_amplitude = f"D{11 * i + 10}"

        if beamName == 'a':  
            ws[BeamA_rel_amplitude] = 1
            ws[BeamB_rel_amplitude] = 0

        elif beamName == 'b':  
            ws[BeamA_rel_amplitude] = 0
            ws[BeamB_rel_amplitude] = 1

        global_amplitude = f"D{11 * i + 11}"  
        ws[global_amplitude] = 1
        
        BeamA_label = f"A{11 * i + 2}"
        ws[BeamA_label] = 'Beam A'
        BeamA_label = f"A{11 * i + 3}"
        ws[BeamA_label] = 'Beam A'
        
        tiltX_label = f"C{11 * i + 2}"
        ws[tiltX_label] = 'Grating period X (plm pix)'
        tiltY_label = f"C{11 * i + 3}"
        ws[tiltY_label] = 'Grating period Y (plm pix)'
        
        BeamB_label = f"A{11 * i + 4}"
        ws[BeamB_label] = 'Beam B'
        BeamB_label = f"A{11 * i + 5}"
        ws[BeamB_label] = 'Beam B'
        
        tiltX_label = f"C{11 * i + 4}"
        ws[tiltX_label] = 'Grating period X (plm pix)'
        tiltY_label = f"C{11 * i + 5}"
        ws[tiltY_label] = 'Grating period Y (plm pix)'
        
        BeamA_label = f"A{11 * i + 6}"
        ws[BeamA_label] = 'Beam A'
        phase_label = f"C{11 * i + 6}"
        ws[phase_label] = 'Relative phase (%)'
        
        BeamB_label = f"A{11 * i + 7}"
        ws[BeamB_label] = 'Beam B'
        phase_label = f"C{11 * i + 7}"
        ws[phase_label] = 'Relative phase (%)'
        
        Beam_label = f"A{11 * i + 8}"
        ws[Beam_label] = 'Beam'
        phase_label = f"C{11 * i + 8}"
        ws[phase_label] = 'Global phase (%)'
        
        BeamA_label = f"A{11 * i + 9}"
        ws[BeamA_label] = 'Beam A'
        amp_label = f"C{11 * i + 9}"
        ws[amp_label] = 'Relative amplitude'
        
        BeamB_label = f"A{11 * i + 10}"
        ws[BeamB_label] = 'Beam B'
        amp_label = f"C{11 * i + 10}"
        ws[amp_label] = 'Relative amplitude'
        
        Beam_label = f"A{11 * i + 11}"
        ws[Beam_label] = 'Beam'
        amp_label = f"C{11 * i + 11}"
        ws[amp_label] = 'Global amplitude'
        
       

    # Save the workbook
    wb.save(full_filename)
    print(f"\n Wrote {len(commands)} x-y command pairs to '{full_filename}'")
    #%%
    from matplotlib.patches import Circle
    import matplotlib
    from matplotlib import cm
    from matplotlib.colors import Normalize

    plt.close('all')
    img = np.zeros((512 , 512))
    fig,ax = plt.subplots(1)
    ax.set_aspect('equal')
    ax.imshow(img , cmap = 'gray')

    # Normalize the indices for colormap
    norm = Normalize(vmin=0, vmax=len(centroid_x)-1)
    cmap = matplotlib.colormaps['jet']

    # Add circles with colors from the colormap
    for i, (xx, yy) in enumerate(zip(centroid_x, centroid_y)):
        color = cmap(norm(i))  # RGBA tuple from colormap
        circ = Circle((xx, yy), 3, facecolor=color, edgecolor='black')
        ax.add_patch(circ)
        ax.set_title('Measured centroid positions from camera')

    
    fig_filename = os.path.join(time_folder , beam_name + '.png') 
    plt.savefig(fig_filename, dpi=300, bbox_inches='tight')  # dpi/bbox optional but recommended

    #plt.show()

    print('\nSwitching back to free streaming')
    self.camera.TriggerMode.SetValue("Off")        
    offset_x = 0
    offset_y = 0        
    self.camera.OffsetX.SetValue(offset_x)
    self.camera.OffsetY.SetValue(offset_y)
    image_height = 512
    image_width = 512
    self.camera.Width.SetValue(image_width)
    self.camera.Height.SetValue(image_height)
    self.camera.StartGrabbing(pylon.GrabStrategy_LatestImageOnly)
    plm.set_frame(0)
    time.sleep(0.5)
    plm.play()
    plm.play()
